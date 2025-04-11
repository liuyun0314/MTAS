import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
import torch
import random
import simpy
import copy
import datetime
import time
import argparse
import openpyxl
import numpy as np
from params import args
import matplotlib.pyplot as plt
# from Algorithms.MATMS2 import MATMS
from Algorithms.MATMS import MATMS
from Algorithms.heuristics import heuristics
from Environments.workShop import JobShop
from peerCompitors.AMDQN.models import Actor_obj, Actor_rule
from Algorithms.AMDQN import AMDQN
from Algorithms.PPO import PPO
from Algorithms.HMAPPO import HMAPPO
from Algorithms.DMDDQN import DMDDQN
from peerCompitors.HMAPPO.HMAPPO_models import *
from peerCompitors.PPO.model_PPO import Actor
from peerCompitors.DMDDQN.DMDDQN_model import *
# from Runner.Transformer_policy import TransformerPolicy as Policy
# from Runner.transformer_policy_Ddecoder2 import TransformerPolicy as Policy
from Runner.Transformer_policy import TransformerPolicy as Policy

class OverallExperiment:
    def __init__(self, args):
        self.iter = args.iter
        self.num_new_jobs = args.num_new_jobs
        self.num_machines = args.num_machines
        self.number_of_testinstances = args.number_of_testinstances
        self.obs_dim = args.obs_dim  # 一个智能体的观察空间维度
        self.share_obs_dim = args.share_obs_dim  # 所有智能体的共享观察空间维度
        self.reasoning_dim = args.reasoning_dim  # 一个智能体的推理维度
        self.act_dim = args.n_action
        self.act_num = 1
        self.device = args.device
        self.num_agents = args.num_tasks + 1
        self.use_centralized_V = args.use_centralized_V
        self.root_dir = '/home/yunliu/Code/3/results/Records/'
        self.test_plots_dir = '/home/yunliu/Code/3/results/test_plots/'
        # self.policy_dir = '/home/yunliu/Code/3/Models/Ddecoder/strainJSP4/7_3_16_43/transformer_620.pt'
        # self.policy_dir = '/home/yunliu/Code/3/Models/Ddecoder/2/7_10_19_28/transformer_990.pt'
        self.policy_dir = '/home/yunliu/Code/3/trainingResults/Sdecoder/0.95/3_4_9_13/Runner_model/transformer_1266218.pt'
        # self.policy_dir = '/home/yunliu/Code/3/trainingResults/Sdecoder/0.95/3_8_14_50/Runner_model/transformer_1630.pt'
        # self.policy_dir = '/home/yunliu/Code/3/Models/Ddecoder/strainJSP4/7_3_19_50/transformer_125618.pt'
        #  /home/yunliu/Code/3/Models/6_11_15_19/transformer_1870.pt  /home/yunliu/Code/3/Models/6_12_19_5/transformer_138785.pt    /home/yunliu/Code/3/Models/6_13_15_43/transformer_62142.pt    /home/yunliu/Code/3/Models/6_14_9_2/transformer_123287.pt
        # /home/yunliu/Code/3/Models/6_19_14_3/transformer_30676.pt
        # #/home/yunliu/Code/3/Models/6_17_17_30/transformer_1236283.pt
        # /home/yunliu/Code/3/Models/6_18_19_36/transformer_1835999.pt
        # /home/yunliu/Code/3/Models/6_18_8_59/transformer_182305.pt
        # /home/yunliu/Code/3/Models/6_18_8_59/transformer_1172487.pt
        # /home/yunliu/Code/3/Models/6_18_8_59/transformer_1845847.pt
        # /home/yunliu/Code/3/Models/6_20_21_0/transformer_6063.pt
        now = datetime.datetime.now()
        self.now_time = str(now.month) + '_' + str(now.day) + '_' + str(now.hour) + '_' + str(now.minute)

    def test(self, algorithms, u, scale):
    # def test(self, algorithms, config):
        args.num_new_jobs = scale
        args.E_utliz = u
        save_dir = '/home/yunliu/Code/3/results/RulesRecord/' + str(args.num_new_jobs) + '_' + str(self.now_time) + '.xlsx'
        all_objectives = np.zeros((self.number_of_testinstances, len(algorithms), 5))
        all_rewards = np.zeros((self.number_of_testinstances, len(algorithms)))
        all_timeCosts = np.zeros((self.number_of_testinstances, len(algorithms)))
        # self.number_of_testinstances = config.number_of_testinstances
        seeds = random.sample(range(0, 1000000), self.number_of_testinstances)
        counter = -1
        # args.num_new_jobs = config.num_new_jobs

        # args.device = config.device
        # args.E_utliz = config.E_utliz
        wb = openpyxl.Workbook()
        ws = wb.active

        policy = Policy(args, self.num_agents, self.device)
        policy.transformer.load_state_dict(torch.load(self.policy_dir))
        for seed in seeds:
            counter += 1
            test_sim = self.make_test_env(args, seed)
            alg_objectives = np.full((len(algorithms), 4), np.inf)
            alg_test_reward = np.full((len(algorithms)), -np.inf)
            all_objs = np.zeros((len(algorithms),5))
            for algorithm in algorithms:
                sim = copy.deepcopy(test_sim)
                if algorithm == 'MATMS':
                    sht = wb.create_sheet()
                    sht.title = 'Instance' + str(counter)
                    sht.cell(1, 1, 'Task1_RA')
                    sht.cell(1, 2, 'Task2_RA')
                    sht.cell(1, 3, 'Task3_RA')
                    sht.cell(1, 4, 'SA')
                    # args.reasoning_dim = config.reasoning_dim
                    # args.n_block = config.n_block
                    # args.n_embd = config.n_embd
                    # args.n_head = config.n_head
                    wb.save(save_dir)
                    t1 = time.time()
                    test_reward, objectives = MATMS(sim, seed, policy, sht)
                    t2 = time.time()
                else:
                    if algorithm == 'AMDQN':
                        policy_obj = Actor_obj(10)
                        policy_rule = Actor_rule(50)
                        policy_obj.load_state_dict(torch.load('/home/yunliu/Code/3/Models/AMDQN/7_5_19_56/obj_400.pt'))
                        policy_rule.load_state_dict(torch.load('/home/yunliu/Code/3/Models/AMDQN/7_5_19_56/rule_400.pt'))
                        t1 = time.time()
                        test_reward, objectives = AMDQN(sim, policy_obj, policy_rule)
                        t2 = time.time()
                    else:
                        if algorithm == 'DMDDQN':
                            Actor_objective = Actor_obj_DMDDQN(10)
                            Actor_rules = Actor_rule_DMDDQN(50)
                            Actor_objective.load_state_dict(torch.load('/home/yunliu/Code/3/Models/DMDDQN/7_7_16_50/obj_400.pt'))
                            Actor_rules.load_state_dict(torch.load('/home/yunliu/Code/3/Models/DMDDQN/7_7_16_50/rule_400.pt'))
                            t1 = time.time()
                            DMDDQN(sim, Actor_objective, Actor_rules)
                            t2 = time.time()
                        else:
                            if algorithm == 'PPO':
                                ploicy = Actor
                                ploicy.load_state_dict(torch.load('/home/yunliu/Code/3/Models/PPO/7_5_22_25/PPO_400.pt'))
                                t1 = time.time()
                                test_reward, objectives = PPO(sim, ploicy)
                                t2 = time.time()
                            else:
                                if algorithm == 'HMAPPO':
                                    Hid_Size = 200
                                    actor_obj = Actor_objAgent(Hid_Size)
                                    actor_routing = Actor_routing(Hid_Size)
                                    actor_sequencing = Actor_sequencing(Hid_Size)
                                    actor_obj.load_state_dict(torch.load('/home/yunliu/Code/3/Models/HMAPPO/7_6_16_53/HMAPPO_obj1000.pt'))
                                    actor_routing.load_state_dict(torch.load('/home/yunliu/Code/3/Models/HMAPPO/7_6_16_53/HMAPPO_routing1000.pt'))
                                    actor_sequencing.load_state_dict(torch.load('/home/yunliu/Code/3/Models/HMAPPO/7_6_16_53/HMAPPO_sequecing1000.pt'))
                                    t1 = time.time()
                                    test_reward, objectives = HMAPPO(sim, actor_routing, actor_sequencing, actor_obj)
                                    t2 = time.time()
                                else:
                                    if algorithm == 'Random':
                                        routings = ['SPTM', 'NINQ', 'WINQ', 'LWT']
                                        sequencings = ['SPT', 'LPT', 'MWKR', 'EDD', 'MOPNR']
                                        tmp_reward = np.zeros(self.iter)
                                        tmp_objectives = np.zeros((self.iter, 4))
                                        t1 = time.time()
                                        for i in range(self.iter):
                                            tmp_sim = copy.deepcopy(sim)
                                            r1 = random.choice(routings)
                                            r2 = random.choice(sequencings)
                                            test_reward, objectives = heuristics(r1, r2, tmp_sim)
                                            tmp_reward[i] = test_reward
                                            tmp_objectives[i, :] = objectives
                                        test_reward = np.mean(tmp_reward)
                                        objectives = np.mean(tmp_objectives, axis=0)
                                        t2 = time.time()
                                    else:
                                        rules = algorithm.split('_')
                                        t1 = time.time()
                                        test_reward, objectives = heuristics(rules[0], rules[1], sim)
                                        t2 = time.time()
                all_objs[algorithms.index(algorithm), :4] = objectives
                all_objs[algorithms.index(algorithm), 4] = t2 - t1
                alg_objectives[algorithms.index(algorithm), :] = objectives
                alg_test_reward[algorithms.index(algorithm)] = test_reward
            wb.close()
            all_objectives[counter, :, :] = all_objs
            # all_objectives[counter, :, :] = alg_objectives
            all_rewards[counter, :] = alg_test_reward
        # print(
        #     "eval average episode rewards: {:.2f} - > WTmean: {:.2f}, WFmean: {:.2f}, WTmax: {:.2f}, machine_UR: {:.2f}, span: {:.2f}.".format(
        #         round(train_episode_rewards, 2), round(WTmean, 2), round(WFmean, 2), round(WTmax, 2),
        #         round(machine_UR, 2),
        #         round(span, 2)))
        self.plot_record(algorithms, all_rewards, all_objectives)
        self.record(algorithms, all_rewards, all_objectives)

    def warmup(self, seed, sim):
        obs, share_obs, ava = sim.reset(seed, sim.arrival_interval)
        if self.use_centralized_V:
            # share_obs = obs
            share_obs = np.concatenate(obs)
        return obs, share_obs, ava

    def make_test_env(self, args, seed):
        env = simpy.Environment()
        # Create a Workshop object
        sim = JobShop(env, args, seed)
        # sim.starTime = np.random.randint(1, 2000)
        # sim.decision_points.append(sim.starTime)
        # repair_time = np.random.randint(args.repair_time_range[0], args.repair_time_range[1])
        # sim.failure_time = sim.starTime + repair_time
        # sim.decision_points.append(sim.failure_time)
        # sim.decision_points = sorted(sim.decision_points)
        return sim

    def record(self, algorithms, test_rewards, objectives):

        # now = datetime.datetime.now()
        # now_time = str(now.month) + '_' + str(now.day) + '_' + str(now.hour) + '_' + str(now.minute)
        save_dir = self.root_dir + self.now_time + '/'
        y_labels = ['WT_mean', 'WF_mean', 'WT_max', 'machine_UR', 'timeCost']
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
        xls = openpyxl.Workbook()
        # record objectives
        for agent in range(self.num_agents+1):
            sht = xls.create_sheet()
            sht.title = y_labels[agent]
            sht.cell(1, 1, 'Algorithm')
            sht.cell(1, self.number_of_testinstances + 2, 'avg_obj')
            sht.cell(1, self.number_of_testinstances + 3, 'Rankings')
            avg_obj = np.mean(objectives[:, :, agent], axis=0)
            rankings = np.argsort(avg_obj)

            for i in range(self.number_of_testinstances):
                sht.cell(1, i+2, 'Instance' + str(i))
                for j in range(len(algorithms)):
                    sht.cell(j+2, 1, algorithms[j])
                    sht.cell(j+2, i+2, objectives[i, j, agent])

            for i in range(len(algorithms)):
                sht.cell(i+2, self.number_of_testinstances + 2, avg_obj[i])
                sht.cell(i + 2, self.number_of_testinstances + 3, rankings[i])

        # record test rewards
        avg_rewards = np.mean(test_rewards, axis=0)
        rankings = np.argsort(avg_rewards)[::-1]
        sht = xls.create_sheet()
        sht.title = 'Test_Rewards'
        sht.cell(1, 1, 'Algorithm')
        sht.cell(1, 2, 'Test_Reward')
        for a in range(self.number_of_testinstances):
            sht.cell(1, a+2, 'Instance' + str(a))
            for i in range(len(algorithms)):
                sht.cell(i+2, 1, algorithms[i])
                sht.cell(i+2, a+2, test_rewards[a, i])
        for i in range(len(algorithms)):
            sht.cell(i + 2, self.number_of_testinstances + 2, rankings[i])

        xls_path = save_dir + str(self.num_new_jobs) + '*' + str(self.num_machines) + '_Record.xlsx'
        xls.save(xls_path)
        # xls.save(save_dir + 'Record.xlsx')

    def plot_record(self, algorithms, test_rewards, objectives):

        colors = plt.cm.Set3(np.linspace(0, 1, len(algorithms)))

        # plot test objectives
        fig, axs = plt.subplots(2, 2, figsize=(12, 8))
        X = np.arange(self.number_of_testinstances)
        y_labels = ['WT_mean', 'WT_max', 'WF_mean', 'machine_UR']
        counter = 0
        for i in range(2):
            for j in range(2):
                objs = objectives[:, :, counter]     # 获取第counter个目标的所有实例的结果
                axs[i, j].set_xlabel('Instance')
                axs[i, j].set_ylabel(y_labels[counter])
                axs[i, j].set_title(f"{y_labels[counter]}")
                for k in range(len(algorithms)):
                    if k == 0:
                        axs[i, j].plot(X, objs[:, k], label=algorithms[k], color='red', linewidth=1.5)
                    else:
                        axs[i, j].plot(X, objs[:, k], label=algorithms[k], color=colors[k-1], linewidth=1.5)
                # axs[i, j].legend()  # 显示图例
                counter += 1
        plt.tight_layout()
        sub_figure_file = self.test_plots_dir + self.now_time + '/'
        if not os.path.exists(sub_figure_file):
            os.makedirs(sub_figure_file)
        # figure_file = sub_figure_file + 'objectives.png'
        # plt.savefig(figure_file)
        figure_file = sub_figure_file + str(self.num_new_jobs) + '*' + str(self.num_machines) + '_objectives.png'
        plt.savefig(figure_file)
        # plt.show()

        plt.figure(figsize=(10, 6))
        for i in range(len(algorithms)):
            if i == 0:
                plt.plot(X, test_rewards[:, i], label=algorithms[i], color='red', linewidth=1.5)
            else:
                plt.plot(X, test_rewards[:, i], label=algorithms[i], color=colors[i-1], linewidth=1.5)
        plt.xlabel('Instances')
        plt.ylabel('test_rewards')
        plt.title('Algorithm Solutions Over instances')
        plt.legend()
        # figure_file = sub_figure_file + 'test_rewards.png'
        # plt.savefig(figure_file)
        figure_file = sub_figure_file + str(self.num_new_jobs) + '*' + str(self.num_machines) + '_test_rewards.png'
        plt.savefig(figure_file)
        # plt.show()

if __name__ == '__main__':
    parser = argparse.ArgumentParser("Hyperparameter Setting for PPO-discrete")
    parser.add_argument("--reasoning_dim", type=int, default=64, help="dim of the reasoning")
    parser.add_argument("--n_block", type=int, default=4)  # 2
    parser.add_argument("--n_embd", type=int, default=128)   # 64
    parser.add_argument("--n_head", type=int, default=1)
    parser.add_argument("--num_new_jobs", type=int, default=100, help="The number of new jobs (default: 5000)")
    parser.add_argument("--number_of_testinstances", type=int, default=30, help="The number of scheduling instances")
    parser.add_argument("--device", type=str, default='cuda:2', help="The device to run the code")
    parser.add_argument("--E_utliz", type=float, default=0.95, help="The machine utilization ")

    config = parser.parse_args()

    # routing = ['SPTM', 'NINQ', 'WINQ', 'LWT']
    # sequencing = ['SPT', 'LPT', 'MWKR', 'EDD', 'MOPNR']
    # algorithms = ['MATMS']
    # algorithms = ['MATMS', 'Random', 'SPTM_LPT']   #  , 'SPTM_SPT' , 'SPTM_EDD', 'SPTM_MWKR', 'SPTM_MOPNR'
    # algorithms = ['MATMS', 'Random', 'SPTM_LPT', 'SPTM_SPT' , 'SPTM_EDD', 'SPTM_MWKR', 'SPTM_MOPNR']
    # algorithms = ['MATMS', 'AMDQN', 'SPTM_LPT', 'SPTM_SPT' , 'SPTM_EDD', 'SPTM_MWKR', 'SPTM_MOPNR']
    # algorithms = ['MATMS', 'HMAPPO', 'PPO', 'DMDDQN', 'AMDQN', 'SPTM_LPT', 'SPTM_SPT', 'SPTM_EDD', 'SPTM_MWKR', 'SPTM_MOPNR']
    # algorithms = ['MATMS', 'HMAPPO', 'PPO', 'DMDDQN', 'AMDQN',  'Random', 'SPTM_SPT', 'SPTM_EDD', 'SPTM_MWKR']
    # algorithms = ['MATMS', 'HMAPPO', 'PPO', 'DMDDQN', 'AMDQN', 'Random', 'LWT_MOPNR', 'LWT_MWKR', 'LWT_SPT', 'LWT_LPT', 'SPTM_LPT', 'SPTM_SPT', 'SPTM_MWKR', 'SPTM_MOPNR']
    algorithms = ['MATMS', 'HMAPPO', 'PPO', 'DMDDQN', 'AMDQN', 'SPTM_SPT']
    # algorithms = ['MATMS', 'HMAPPO', 'PPO', 'DMDDQN', 'AMDQN',  'Random']
    # algorithms = ['MATMS', 'SPTM_LPT', 'SPTM_SPT', 'SPTM_EDD', 'SPTM_MWKR', 'SPTM_MOPNR']
    # hybrid_rules = [rule + '_' + prefix for rule in routing for prefix in sequencing]
    # algorithms.extend(hybrid_rules)
    experiment = OverallExperiment(args)
    # experiment.test(algorithms)
    # experiment.test(algorithms, config)
    scale = 100
    U = [0.95, 0.85, 0.75]
    for u in U:
        experiment.E_utliz = u
        experiment.test(algorithms, u, scale)