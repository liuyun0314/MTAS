import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
import simpy
import torch
import random
import socket
import datetime
import openpyxl
import numpy as np
from params import args
from pathlib import Path
from Runner.Runner import Runner
import matplotlib.pyplot as plt
from Environments.workShop import JobShop

def make_train_env(args):
    env = simpy.Environment()
    seed = random.randint(0, 1000000)
    # Create a Workshop object
    sim = JobShop(env, args, seed)
    # sim.starTime = np.random.randint(1, 2000)
    # sim.decision_points.append(sim.starTime)
    # repair_time = np.random.randint(args.repair_time_range[0], args.repair_time_range[1])
    # sim.failure_time = sim.starTime + repair_time
    # sim.decision_points.append(sim.failure_time)
    # sim.decision_points = sorted(sim.decision_points)
    return sim, seed

def make_eval_env(args):
    env = simpy.Environment()
    # Create a Workshop object
    sim = JobShop(env, args, 50000)
    # sim.starTime = np.random.randint(1, 2000)
    # sim.decision_points.append(sim.starTime)
    # repair_time = np.random.randint(args.repair_time_range[0], args.repair_time_range[1])
    # sim.failure_time = sim.starTime + repair_time
    # sim.decision_points.append(sim.failure_time)
    # sim.decision_points = sorted(sim.decision_points)
    return sim

def plot_graph(done_episodes_rewards, policy_loss, eval_results, save_path):

    intrval = 100

    # root_figure_file = '/home/yunliu/Code/3/results/MATMS/trainJSP/'
    # now = datetime.datetime.now()
    # now_time = str(now.month) + '_' + str(now.day) + '_' + str(now.hour) + '_' + str(now.minute)

    ## plot training records
    y_labels = ['training_reward', 'eval_reward']
    fig, axs = plt.subplots(1, 2, figsize=(12, 4))
    X = range(len(done_episodes_rewards))
    axs[0].set_xlabel('episode')
    axs[0].set_ylabel('training_reward')
    axs[0].set_title(f"{y_labels[0]}")
    axs[0].plot(X, done_episodes_rewards)
    axs[0].set_xticks(X[::intrval])


    X = range(len(policy_loss))
    axs[1].set_xlabel('episode')
    axs[1].set_ylabel('policy_loss')
    axs[1].set_title(f"{y_labels[1]}")
    axs[1].plot(X, policy_loss)
    axs[1].set_xticks(X[::intrval])
    plt.tight_layout()
    # sub_figure_file = root_figure_file + 'traning_plot/'
    # if not os.path.exists(sub_figure_file):
    #     os.makedirs(sub_figure_file)
    figure_file = save_path + str(args.n_block) + '_' + str(args.num_new_jobs) + '_reward.png'
    # figure_file = sub_figure_file + now_time + '_trainingData.png'
    plt.savefig(figure_file)
    plt.show()

    ## plot objectives
    y_labels = ['WT_mean', 'WT_max', 'WF_mean', 'machine_UR']
    fig, axs = plt.subplots(2, 2, figsize=(12, 8))
    results = np.array(eval_results)
    count = 0
    for i in range(2):
        for j in range(2):
            X = range(len(results[:, count]))
            axs[i, j].set_xlabel('episode')
            axs[i, j].set_ylabel(y_labels[count])
            axs[i, j].set_title(f"{y_labels[count]}")
            axs[i, j].plot(X, results[:, count])
            axs[i, j].set_xticks(X[::intrval])
            count += 1
    plt.tight_layout()
    # sub_figure_file = root_figure_file + 'objectives/'
    # if not os.path.exists(sub_figure_file):
    #     os.makedirs(sub_figure_file)
    figure_file = save_path + str(args.n_block) + '_' + str(args.num_new_jobs) + '_objectives.png'
    plt.savefig(figure_file)

def training_records(done_episodes_rewards, policy_loss, eval_results, save_dir):
    xls = openpyxl.Workbook()
    sht = xls.create_sheet()
    sht.title = 'traning_rewards'
    sht.cell(1, 1, 'epoch')
    for i in range(len(done_episodes_rewards)):
        sht.cell(1, i+2, 'epoch'+str(i+1))
        sht.cell(2, i+2, done_episodes_rewards[i])

    sht = xls.create_sheet()
    sht.title = 'policy_loss'
    sht.cell(1, 1, 'epoch')
    for i in range(len(policy_loss)):
        sht.cell(1, i + 2, 'epoch' + str(i + 1))
        sht.cell(2, i + 2, policy_loss[i])

    sht = xls.create_sheet()
    sht.title = 'objectives'
    sht.cell(1, 1, 'epoch')
    sht.cell(2, 1, 'WTmean')
    sht.cell(3, 1, 'WFmean')
    sht.cell(4, 1, 'WTmax')
    sht.cell(5, 1, 'makespan')

    for i in range(len(eval_results)):
        sht.cell(1, i+2, 'epoch'+str(i+1))
        sht.cell(2, i+2, eval_results[i][0])
        sht.cell(3, i+2, eval_results[i][1])
        sht.cell(4, i+2, eval_results[i][2])
        sht.cell(5, i+2, eval_results[i][3])

    xls_name = save_dir + '_' + str(args.n_block) + '_' + str(args.num_new_jobs) + '_traning_Record.xlsx'
    xls.save(xls_name)

def record_rewards(all_rewards, save_dir):
    xls = openpyxl.Workbook()
    sht = xls.create_sheet()
    sht.title = 'traning_rewards'
    sht.cell(1, 1, 'epoch')
    for i, rewards in enumerate(all_rewards):
        for j in range(len(rewards)):
            sht.cell(1, j+2, 'epoch'+str(j+1))
            sht.cell(i + 2, 1, str(i+1))
            sht.cell(i+2, j+2, rewards[j])

    xls_name = save_dir + '_' + str(args.n_block) + '_' + str(args.num_new_jobs) + '_RecordRewards.xlsx'
    xls.save(xls_name)


def main(args):
    import time
    now = datetime.datetime.now()
    now_time = str(now.month) + '_' + str(now.day) + '_' + str(now.hour) + '_' + str(now.minute)
    root_path = '/home/yunliu/Code/3/trainingResults/Sdecoder/' + str(args.E_utliz) + '/'
    save_path = root_path + now_time + '/'
    if not os.path.exists(save_path):
        os.makedirs(save_path)
    #device
    if args.cuda and torch.cuda.is_available():
        print("choose to use gpu...")
        device = torch.device("cuda:0")
        torch.set_num_threads(args.n_training_threads)
        if args.cuda_deterministic:
            torch.backends.cudnn.benchmark = False
            torch.backends.cudnn.deterministic = True
    else:
        print("choose to use cpu...")
        device = torch.device("cpu")
        torch.set_num_threads(args.n_training_threads)

    # env
    envs, seed = make_train_env(args)
    eval_envs = make_eval_env(args)
    runner = Runner(args, envs, eval_envs, seed, save_path)

    start_time = time.time()
    training_reward, training_loss, eval_results, all_rewards = runner.run()
    end_time = time.time()
    print("Training Time Cost: ", end_time - start_time)

    # plot
    plot_graph(training_reward, training_loss, eval_results, save_path)
    training_records(training_loss, training_loss, eval_results, save_path)
    record_rewards(all_rewards, save_path)

if __name__ == '__main__':
    main(args)
