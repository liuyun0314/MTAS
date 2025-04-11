import os
import time
import torch
import numpy as np
import copy
import datetime
import openpyxl
import random
from tqdm import tqdm
from MAT.utils.shared_buffer import SharedReplayBuffer
from MAT.algorithms.mat.mat_trainer import MATTrainer as TrainAlgo
from Runner.Transformer_policy import TransformerPolicy as Policy
from Environments.Reward import *

def _t2n(x):
    """Convert torch tensor to a numpy array."""
    return x.detach().cpu().numpy()

class Runner(object):
    """
    Base class for training recurrent policies.
    :param config: (dict) Config dictionary containing parameters for training.
    """
    def __init__(self, args, envs, eval_envs, seed, save_path):
        # env = simpy.Environment()
        self.seed = seed
        self.envs = envs
        self.device = args.device
        self.num_agents = args.num_tasks + 1    # num_tasks agents for selecting routing rule for each task and an agent for selecting sequencing rule

        # parameters
        self.algorithm_name = "mat"
        self.use_linear_lr_decay = args.use_linear_lr_decay
        self.hidden_size = args.hidden_size
        self.experiment_name = args.experiment_name
        self.use_centralized_V = args.use_centralized_V
        self.update_interval = args.episode_length
        self.episode_length = 0
        self.max_train_steps = args.max_train_steps
        self.n_rollout_threads = args.n_rollout_threads
        self.eval_episodes = args.eval_episodes
        self.obs_dim = args.obs_dim
        self.share_obs_dim = args.share_obs_dim
        self.save_path = save_path
        # self.xls_name = self.save_path + '_' + str(args.n_block) + '_' + str(args.num_new_jobs) + '_RecordRewards.xlsx'
        # self.xls = openpyxl.Workbook()
        # self.sht = self.xls.create_sheet()
        # self.sht.title = 'traning_rewards'
        # self.sht.cell(1, 1, 'epoch')


        # interval
        self.save_interval = args.save_interval
        self.use_eval = args.use_eval
        self.eval_envs = eval_envs
        self.eval_interval = args.eval_interval

        # policy network
        self.policy = Policy(args, self.num_agents, self.device)
        self.save_dir = args.save_dir

        # algorithm
        self.trainer = TrainAlgo(args, self.policy, self.num_agents, device=self.device)

        # buffer
        self.buffer = SharedReplayBuffer(args, self.num_agents)

        # job shop environment
        self.use_dynamic_reward = args.use_dynamic_reward
        self.machine_queue = {}
        for m in self.envs.machines:
            self.machine_queue[m.name] = []
        # evaluation

    def warmup(self):
        # self.seed = random.randint(0, 1000000)
        self.seed = 100
        obs, share_obs, ava = self.envs.reset(self.seed, self.envs.arrival_interval)
        if self.use_centralized_V:
            # share_obs = obs
            share_obs = np.concatenate(obs)
        self.buffer.share_obs[0] = share_obs.copy()
        self.buffer.obs[0] = obs.copy()
        self.buffer.available_actions[0] = ava.copy()

    def run(self):
        # self.warmup()

        start_time = time.time()

        # episodes = int(self.num_env_steps) // self.episode_length // self.n_rollout_threads
        training_loss = []
        eval_reward = []
        eval_results = []

        done_episodes_rewards = [0]
        all_rewards = []

        train_episode_scores = [0 for _ in range(self.n_rollout_threads)]
        done_episodes_scores = []
        now = datetime.datetime.now()
        now_time = str(now.month) + '_' + str(now.day) + '_' + str(now.hour) + '_' + str(now.minute)
        total_num_steps = -1
        step = -1
        update_times = 0
        optimal_update = 0
        opt_makespan = np.Inf
        opt_WTmean, opt_WFmean, opt_WTmax = 1e3, 1e3, 1e3
        optimal_obj = [opt_WTmean, opt_WFmean, opt_WTmax]
        record_rewards = [[] for _ in range(self.max_train_steps)]

        # for gen in tqdm(range(num_generations), desc="Training Generations", unit="generation"):
        for episode in tqdm(range(self.max_train_steps), desc="Training epochs", unit="epoch"):
            # self.sht.cell(episode + 2, 1, str(episode + 1))
            train_episode_rewards = 0
            train_episode_r = 0
            self.warmup()
            if self.use_linear_lr_decay:
                self.trainer.policy.lr_decay(episode, self.max_train_steps)
            record_step = 0
            while not self.envs.done:
                total_num_steps += 1
                self.envs.new_job_arrival()
                self.envs.machine_failure()
                self.envs.machine_repair()
                if self.envs.env.now == self.envs.decision_points[0]:
                    step += 1
                    # compute return and update network
                    if step == self.update_interval - 1:
                        # if step % self.update_interval == 0 and step != 0:
                        step = -1
                        self.compute()
                        train_info = self.train()
                        training_loss.append(train_info['policy_loss'])
                        if (update_times % self.save_interval == 0):
                            self.save(update_times, now_time)
                        update_times += 1

                    self.envs.decision_points.remove(self.envs.decision_points[0])
                    idle_machines = []
                    for m in self.envs.machines:
                        if m.currentTime <= self.envs.env.now:
                            idle_machines.append(m)
                    if len(self.envs.jobs) == 0 or len(idle_machines) == 0:
                        if len(self.envs.decision_points) == 0:
                            self.envs.env.timeout(1)
                            self.envs.env.run()
                        else:
                            self.envs.env.timeout(self.envs.decision_points[0] - self.envs.env.now)
                            self.envs.env.run()
                        continue
                    values, actions, action_log_probs = self.collect(step)
                    rules = actions.squeeze().tolist()
                    obs, share_obs, r, rewards, dones, available_actions = self.envs.step(rules)
                    #### ly ####
                    rewards_env = np.mean(rewards)
                    share_obs = np.concatenate(obs)

                    train_episode_rewards += np.mean(r)
                    train_episode_r += rewards_env
                    rewards_env = r
                    record_rewards[episode].append(r)
                    self.buffer.insert(share_obs, obs, actions, action_log_probs, values, r, available_actions)
                    # self.sht.cell(1, record_step + 2, 'epoch' + str(record_step + 1))
                    # self.sht.cell(episode + 2, record_step + 2, r)
                    # self.xls.save(self.xls_name)
                    # record_step += 1
                    ##########

                    if len(self.envs.decision_points) == 0:
                        self.envs.env.timeout(1)
                        self.envs.env.run()
                    else:
                        self.envs.env.timeout(self.envs.decision_points[0] - self.envs.env.now)
                        self.envs.env.run()
                if len(self.envs.decision_points) == 0:
                    self.envs.decision_points.append(self.envs.env.now)
                self.envs.decision_points = sorted(self.envs.decision_points)

            # for record_step, r in enumerate(record_rewards[episode]):
            #     record_rewards[episode].append(r)
            #     self.sht.cell(1, record_step + 2, 'epoch' + str(record_step + 1))
            #     self.sht.cell(episode + 2, record_step + 2, r)
            #     self.xls.save(self.xls_name)

            WTmean, WFmean, WTmax, machine_UR, span = self.get_objectives(self.envs.tasks_list, self.envs.machines)
            print("episode {:.2f} : eval average rewards: {:.2f} - > WTmean: {:.2f}, WFmean: {:.2f}, WTmax: {:.2f}, machine_UR: {:.2f}, span: {:.2f}.".format(round(episode, 2),
                    round(train_episode_rewards, 2), round(WTmean, 2), round(WFmean, 2), round(WTmax, 2), round(machine_UR, 2),
                    round(span, 2)))
            eval_result = [WTmean, WFmean, WTmax, span]
            eval_results.append(eval_result)

            if opt_WTmean >= WTmean and opt_WFmean > WFmean and opt_WTmax >= WTmax:
                self.optimal_save(optimal_update, now_time)
                optimal_update += 1
                opt_WTmean = WTmean
                opt_WFmean = WFmean
                opt_WTmax = WTmax
            done_episodes_rewards.append(train_episode_rewards)
            all_rewards.append(train_episode_r)
            train_episode_rewards = 0

        end_time = time.time()
        avg_train_time = (end_time - start_time) / self.max_train_steps
        print("Average training time: {}s".format(avg_train_time))
        self.save(total_num_steps, now_time)
        done_episodes_rewards.remove(done_episodes_rewards[0])
        return done_episodes_rewards, training_loss, eval_results, record_rewards

    def collect(self, step):
        self.trainer.prep_rollout()
        value, action, action_log_prob = self.trainer.policy.get_actions(np.concatenate(self.buffer.share_obs[step]),
                                              np.concatenate(self.buffer.obs[step]),
                                              np.concatenate(self.buffer.available_actions[step]))
        values = np.array(np.split(_t2n(value), self.n_rollout_threads))
        actions = np.array(np.split(_t2n(action), self.n_rollout_threads))
        action_log_probs = np.array(np.split(_t2n(action_log_prob), self.n_rollout_threads))
        return values, actions, action_log_probs

    def insert(self, data):
        obs, share_obs, rewards, dones, available_actions, \
            values, actions, action_log_probs = data

        # original code   由于在我们的环境中，每个task的作业是动态到达的，所以，除非所有的作业全都到达，否则我们无法明确每个任务是否为done，因此，这里我们不使用mask
        dones_env = np.all(dones, axis=1)
        masks = np.ones((self.n_rollout_threads, self.num_agents, 1), dtype=np.float32)
        masks[dones_env == True] = np.zeros(((dones_env == True).sum(), self.num_agents, 1), dtype=np.float32)
        active_masks = np.ones((self.n_rollout_threads, self.num_agents, 1), dtype=np.float32)
        active_masks[dones == True] = np.zeros(((dones == True).sum(), 1), dtype=np.float32)
        active_masks[dones_env == True] = np.ones(((dones_env == True).sum(), self.num_agents, 1), dtype=np.float32)
        if not self.use_centralized_V:
            share_obs = obs
        self.buffer.insert(share_obs, obs, actions, action_log_probs, values, rewards, masks, None, active_masks, available_actions)

    def compute(self):
        """Calculate returns for the collected data."""
        self.trainer.prep_rollout()
        if self.buffer.available_actions is None:
            next_values = self.trainer.policy.get_values(np.concatenate(self.buffer.share_obs[-1]),
                                                         np.concatenate(self.buffer.obs[-1]),
                                                         np.concatenate(self.buffer.masks[-1]))
        else:
            next_values = self.trainer.policy.get_values(np.concatenate(self.buffer.share_obs[-1]),
                                                         np.concatenate(self.buffer.obs[-1]),
                                                         np.concatenate(self.buffer.masks[-1]),
                                                         np.concatenate(self.buffer.available_actions[-1]))
        next_values = np.array(np.split(_t2n(next_values), self.n_rollout_threads))
        self.buffer.compute_returns(next_values, self.trainer.value_normalizer)

    def train(self):
        """Train policies with data in buffer. """
        self.trainer.prep_training()
        train_infos = self.trainer.train(self.buffer)
        self.buffer.after_update()
        return train_infos

    def save(self, episode, now_time):
        """Save policy's actor and critic networks."""
        file_name = self.save_path + 'Runner_model/'
        if not os.path.exists(file_name):
            os.makedirs(file_name)
        self.policy.save(file_name, episode)

    def optimal_save(self, optimal_update, now_time):
        # file_name = self.save_dir + now_time
        file_name = self.save_path + 'Runner_model/'
        if not os.path.exists(file_name):
            os.makedirs(file_name)
        torch.save(self.policy.transformer.state_dict(), str(file_name) + "/optimalTransformer_" + str(optimal_update) + ".pt")


    def eval(self, eval_envs):
        eval_episode = 0
        eval_episode_rewards = []
        one_episode_rewards = [0 for _ in range(self.eval_episodes)]
        eval_episode_scores = []
        one_episode_scores = [0 for _ in range(self.eval_episodes)]

        eval_obs, eval_share_obs, ava = eval_envs.reset(50000, self.eval_envs.arrival_interval)
        while not eval_envs.done:
            eval_envs.new_job_arrival()
            eval_envs.machine_failure()
            eval_envs.machine_repair()
            if eval_envs.env.now == eval_envs.decision_points[0]:
                eval_envs.decision_points.remove(eval_envs.decision_points[0])
                idle_machines = []
                for m in eval_envs.machines:
                    if m.currentTime <= eval_envs.env.now:
                        idle_machines.append(m)
                if len(eval_envs.jobs) == 0 or len(idle_machines) == 0:
                    if len(eval_envs.decision_points) == 0:
                        eval_envs.env.timeout(1)
                        eval_envs.env.run()
                    else:
                        eval_envs.env.timeout(eval_envs.decision_points[0] - eval_envs.env.now)
                        eval_envs.env.run()
                    continue
                self.trainer.prep_rollout()

                #### ly ######
                shared_obs = np.concatenate(eval_share_obs)
                # eval_obs_after = np.concatenate(eval_obs)
                # eval_ava = np.concatenate(ava)
                shared_state = np.broadcast_to(shared_obs, (self.num_agents, self.share_obs_dim))
                broad_eval_obs = np.broadcast_to(eval_obs, (self.num_agents, self.obs_dim))
                broad_ava = np.broadcast_to(ava, (self.num_agents, 9))
                eval_actions = self.trainer.policy.act(shared_state,
                                                       broad_eval_obs,
                                                       broad_ava,
                                                       deterministic=True)
                # eval_actions = self.trainer.policy.act(np.concatenate(eval_share_obs),
                #                             np.concatenate(eval_obs),
                #                             np.concatenate(ava),
                #                             deterministic=True)
                # eval_actions = np.array(np.split(_t2n(eval_actions), self.eval_episodes))
                rules = eval_actions.squeeze().tolist()
                eval_obs, eval_share_obs, _, eval_rewards, eval_dones, ava = eval_envs.step(rules)
                eval_rewards = np.mean(eval_rewards)
                # one_episode_rewards += eval_rewards
                eval_episode += eval_rewards
                # eval_dones_env = np.all(eval_dones, axis=1)

                if len(eval_envs.decision_points) == 0:
                    eval_envs.env.timeout(1)
                    eval_envs.env.run()
                else:
                    eval_envs.env.timeout(eval_envs.decision_points[0] - eval_envs.env.now)
                    eval_envs.env.run()
            if len(eval_envs.decision_points) == 0:
                eval_envs.decision_points.append(eval_envs.env.now)
            eval_envs.decision_points = sorted(eval_envs.decision_points)
        WTmean, WFmean, WTmax, machine_UR, span = self.get_objectives(eval_envs.tasks_list, eval_envs.machines)
        # WTmean = WT_mean_func(eval_envs.tasks_list[0].jobsList)
        # WFmean = WF_mean_func(eval_envs.tasks_list[1].jobsList)
        # WTmax = WT_max_func(eval_envs.tasks_list[2].jobsList)
        # # machine_UR = machine_utilizatioin_ratio(eval_envs.machines)
        # machine_UR = machine_utilizatioin_ratio(eval_envs.machines)

        print("eval average episode rewards: {:.2f} - > WTmean: {:.2f}, WFmean: {:.2f}, WTmax: {:.2f}, machine_UR: {:.2f}, span: {:.2f}.".format(
                round(eval_episode, 2), round(WTmean, 2), round(WFmean, 2), round(WTmax, 2), round(machine_UR, 2), round(span, 2)))

        return eval_episode, [WTmean, WTmax, WFmean, machine_UR]

    def get_objectives(self, tasks_list, machine_list):
        WTmean = WT_mean_func(tasks_list[0].jobsList)
        WFmean = WF_mean_func(tasks_list[1].jobsList)
        WTmax = WT_max_func(tasks_list[2].jobsList)
        machine_UR = machine_utilizatioin_ratio(machine_list)
        pt = [tasks_list[i].endTime for i in range(self.num_agents - 1)]
        span = max(pt)
        return WTmean, WFmean, WTmax, machine_UR, span






